#!/usr/bin/env python3
"""
BLOG XLSX importer (dry-run only).

Назначение:
- Прочитать XLSX-дамп
- Распознать схему (header row или positional fallback)
- Нормализовать поля по docs/architecture/BLOG_CANONICAL_MAPPING.md
- Сформировать отчёт без записи в Sheets/БД

Поле is_publishable в отчёте — по docs/BLOG_CONTRACT_v1.md (is_publishable_row).

Пример:
    python scripts/blog_xlsx_dry_run_importer.py ^
      --xlsx "C:/Users/X230/Downloads/MyWave_Parser_News.xlsx" ^
      --sheet raw_feed ^
      --out-json "reports/blog_xlsx_dry_run_report.json" ^
      --out-md "reports/blog_xlsx_dry_run_report.md"
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

REPO_ROOT = Path(__file__).resolve().parent.parent
import sys

if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from app.services.blog.publishability import PUBLISHABLE_STATUSES_V1, is_publishable_row

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl не установлен. Установите: venv/Scripts/python -m pip install openpyxl"
    ) from exc


KNOWN_HEADERS = {
    "id",
    "source_type",
    "source_name",
    "source_url",
    "created_at",
    "updated_at",
    "published_at",
    "status",
    "published_posts",
    "raw_title",
    "title",
    "summary",
    "lead",
    "final_posts",
    "text",
    "content_md",
    "cover_image_url",
    "image_url",
    "raw_media",
    "raw_tags",
    "tags",
    "slug",
    "lang",
    "checksum",
    "meta_description",
    "seo_description",
}

# positional fallback для случаев, когда первая строка - это данные, а не headers
# (используем только стартовый блок, который нужен для витрины)
POSITIONAL_FALLBACK_HEADERS = [
    "id",
    "source_type",
    "source_name",
    "source_url",
    "created_at",
    "ingest_status",
    "raw_title",
    "raw_content",
    "raw_html",
    "raw_media",
    "lang",
    "raw_tags",
    "checksum",
    "parse_error",
    "debug_info",
    "expert_opinion",
    "need_opinion",
    "updated_at",
    "published_posts",
    "review_queue",
    "status",
    "published_at",
    "final_posts",
    "summary",
    "lead",
    "slug",
    "cover_image_url",
    "image_url",
]

STATUS_ENUM_REFERENCE = {"DRAFT", "REVIEW", "READY_TO_PUBLISH", "PUBLISHED", "ARCHIVED"}


def _safe_str(v: Any) -> str:
    return str(v).strip() if v is not None else ""


def _as_bool(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    s = _safe_str(v).lower()
    return s in {"1", "true", "yes", "y", "да"}


def _slugify(title: str, post_id: str) -> str:
    t = (title or "").strip().lower().replace("ё", "е")
    t = "".join(ch if ch.isalnum() or ch in ("-", "_") else "-" for ch in t)
    t = "-".join([p for p in t.split("-") if p]) or "post"
    short = hashlib.md5((post_id or "").encode("utf-8")).hexdigest()[:6]
    return f"{t}-{short}"


def _extract_cover(raw: Dict[str, Any]) -> str:
    cover = _safe_str(raw.get("cover_image_url") or raw.get("image_url"))
    if cover:
        return cover
    raw_media = _safe_str(raw.get("raw_media"))
    if not raw_media:
        return ""
    if raw_media.startswith("[") and raw_media.endswith("]"):
        try:
            arr = json.loads(raw_media)
            if isinstance(arr, list) and arr:
                return _safe_str(arr[0])
        except Exception:
            return ""
    if raw_media.startswith("http"):
        return raw_media
    return ""


def _row_header_score(values: List[str]) -> int:
    normalized = {_safe_str(v).lower() for v in values if _safe_str(v)}
    return len(normalized.intersection(KNOWN_HEADERS))


def detect_schema(rows: List[List[Any]]) -> Tuple[bool, Optional[int], List[str], Dict[str, Any]]:
    best_idx = None
    best_score = -1
    max_scan = min(len(rows), 30)
    for i in range(max_scan):
        vals = [_safe_str(v) for v in rows[i]]
        score = _row_header_score(vals)
        if score > best_score:
            best_score = score
            best_idx = i

    # если нашли хотя бы несколько канонических заголовков - считаем это header row
    if best_idx is not None and best_score >= 4:
        headers = [_safe_str(v) for v in rows[best_idx]]
        return True, best_idx, headers, {"score": best_score, "mode": "header_row"}

    # иначе positional fallback
    max_cols = max((len(r) for r in rows), default=len(POSITIONAL_FALLBACK_HEADERS))
    headers = POSITIONAL_FALLBACK_HEADERS + [
        f"extra_col_{i}" for i in range(len(POSITIONAL_FALLBACK_HEADERS), max_cols)
    ]
    return False, None, headers, {"score": best_score, "mode": "positional_fallback"}


def normalize_row(raw: Dict[str, Any]) -> Dict[str, Any]:
    post_id = _safe_str(raw.get("id") or raw.get("news_id") or raw.get("raw_id"))
    title = _safe_str(raw.get("title") or raw.get("raw_title"))
    excerpt = _safe_str(raw.get("summary") or raw.get("lead"))
    content_md = _safe_str(raw.get("final_posts") or raw.get("text") or raw.get("content_md"))
    status = _safe_str(raw.get("status"))
    published_posts = _as_bool(raw.get("published_posts"))
    slug = _safe_str(raw.get("slug")) or _slugify(title or "post", post_id or "unknown")
    pub_v1 = is_publishable_row(raw)

    return {
        "id": post_id,
        "title": title,
        "slug": slug,
        "excerpt": excerpt,
        "content_md": content_md,
        "content_html": "[generated at runtime]",
        "cover_image_url": _extract_cover(raw),
        "tags": _safe_str(raw.get("raw_tags") or raw.get("tags")),
        "status": status,
        "published_at": _safe_str(raw.get("published_at")),
        "updated_at": _safe_str(raw.get("updated_at")),
        "source_name": _safe_str(raw.get("source_name")),
        "source_url": _safe_str(raw.get("source_url")),
        "is_publishable": pub_v1,
        "published_posts": published_posts,
        "legacy_ingest_status": _safe_str(raw.get("ingest_status")),
    }


def to_markdown(report: Dict[str, Any]) -> str:
    s = report["summary"]
    schema = report["schema"]
    lines = [
        "# BLOG XLSX Dry-Run Report",
        "",
        "## Input",
        f"- xlsx: `{report['input']['xlsx_path']}`",
        f"- sheet: `{report['input']['sheet']}`",
        "",
        "## Schema detection",
        f"- mode: `{schema['mode']}`",
        f"- header_row_index: `{schema['header_row_index']}`",
        f"- header_score: `{schema['score']}`",
        "",
        "## Summary",
        f"- total_rows_scanned: `{s['total_rows_scanned']}`",
        f"- valid_rows: `{s['valid_rows']}`",
        f"- invalid_rows: `{s['invalid_rows']}`",
        f"- missing_title: `{s['missing_title']}`",
        f"- missing_slug: `{s['missing_slug']}`",
        f"- missing_content: `{s['missing_content']}`",
        f"- potential_publishable: `{s['potential_publishable']}`",
        "",
        "## Status distribution (top)",
    ]
    for name, cnt in report["status_distribution"][:15]:
        lines.append(f"- `{name}`: `{cnt}`")

    lines += [
        "",
        "## Conflicts",
        f"- unknown_editorial_status_count: `{len(report['unknown_editorial_status_rows'])}`",
        f"- ingest_ok_with_non_publishable_count: `{len(report['ingest_ok_non_publishable_rows'])}`",
        f"- approved_status_count: `{len(report['approved_status_rows'])}`",
        "",
        "## Notes",
        "- Report is dry-run only. No writes to Sheets/DB were performed.",
    ]
    return "\n".join(lines) + "\n"


def main() -> int:
    parser = argparse.ArgumentParser(description="Dry-run importer audit for blog XLSX")
    parser.add_argument("--xlsx", required=True, help="Path to XLSX file")
    parser.add_argument("--sheet", default=None, help="Worksheet name (default: active)")
    parser.add_argument("--out-json", default="reports/blog_xlsx_dry_run_report.json")
    parser.add_argument("--out-md", default="reports/blog_xlsx_dry_run_report.md")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx).expanduser().resolve()
    if not xlsx_path.exists():
        raise SystemExit(f"Файл не найден: {xlsx_path}")

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    if args.sheet:
        if args.sheet not in wb.sheetnames:
            raise SystemExit(f"Лист '{args.sheet}' не найден. Доступно: {wb.sheetnames}")
        ws = wb[args.sheet]
    else:
        ws = wb.active

    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    has_header, header_idx, headers, meta = detect_schema(rows)

    start_idx = (header_idx + 1) if has_header and header_idx is not None else 0
    data_rows = rows[start_idx:]

    normalized_rows: List[Dict[str, Any]] = []
    invalid_row_indices: List[int] = []
    missing_title = 0
    missing_slug = 0
    missing_content = 0

    status_counter: Counter[str] = Counter()
    unknown_editorial_status_rows: List[int] = []
    ingest_ok_non_publishable_rows: List[int] = []
    approved_status_rows: List[int] = []

    for i, row in enumerate(data_rows, start=start_idx + 1):
        raw = {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}
        post = normalize_row(raw)
        normalized_rows.append(post)

        status = _safe_str(post["status"])
        status_counter[status or "(empty)"] += 1

        if not post["title"]:
            missing_title += 1
        if not post["slug"]:
            missing_slug += 1
        if not post["content_md"]:
            missing_content += 1

        row_invalid = False
        if not post["title"]:
            row_invalid = True
        if not post["slug"]:
            row_invalid = True
        if row_invalid:
            invalid_row_indices.append(i)

        status_upper = status.upper()
        if status and status_upper not in STATUS_ENUM_REFERENCE and status_upper not in PUBLISHABLE_STATUSES_V1:
            unknown_editorial_status_rows.append(i)
        if status_upper == "APPROVED":
            approved_status_rows.append(i)
        if _safe_str(post["legacy_ingest_status"]).upper() == "OK" and not post["is_publishable"]:
            ingest_ok_non_publishable_rows.append(i)

    valid_rows = len(normalized_rows) - len(invalid_row_indices)
    potential_publishable = sum(1 for p in normalized_rows if p["is_publishable"])

    report: Dict[str, Any] = {
        "input": {"xlsx_path": str(xlsx_path), "sheet": ws.title},
        "schema": {
            "mode": meta["mode"],
            "score": meta["score"],
            "has_header_row": has_header,
            "header_row_index": header_idx,
            "headers_count": len(headers),
            "headers_preview": headers[:40],
        },
        "summary": {
            "total_rows_scanned": len(data_rows),
            "valid_rows": valid_rows,
            "invalid_rows": len(invalid_row_indices),
            "missing_title": missing_title,
            "missing_slug": missing_slug,
            "missing_content": missing_content,
            "potential_publishable": potential_publishable,
        },
        "status_distribution": status_counter.most_common(),
        "unknown_editorial_status_rows": unknown_editorial_status_rows[:200],
        "approved_status_rows": approved_status_rows[:200],
        "ingest_ok_non_publishable_rows": ingest_ok_non_publishable_rows[:200],
        "sample_normalized_rows": normalized_rows[:5],
        "invalid_row_indices": invalid_row_indices[:200],
    }

    out_json = Path(args.out_json)
    out_md = Path(args.out_md)
    out_json.parent.mkdir(parents=True, exist_ok=True)
    out_md.parent.mkdir(parents=True, exist_ok=True)

    out_json.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    out_md.write_text(to_markdown(report), encoding="utf-8")

    print(f"Dry-run report JSON: {out_json}")
    print(f"Dry-run report MD:   {out_md}")
    print(
        "Summary:",
        report["summary"],
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
